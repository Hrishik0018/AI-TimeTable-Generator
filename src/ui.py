# =====================================
# FINAL ui.py (ONE-FILE - NO DUPLICATES)
# =====================================

import os
import threading
import traceback
from datetime import datetime
import pandas as pd
import customtkinter as ctk
from tkinter import filedialog, messagebox

# Try imports from src package or local modules
try:
    from src import data_handler
    from src import scheduler_fix
    from src import scheduler_ortools as ort
    from src import updater_ai
except Exception:
    try:
        import data_handler
    except:
        data_handler = None
    try:
        import scheduler_fix
    except:
        scheduler_fix = None
    try:
        import scheduler_ortools as ort
    except:
        ort = None
    try:
        import updater_ai
    except:
        updater_ai = None


# -----------------------
# Simple CSV writer
# -----------------------
def _write_csv(path, rows):
    import csv
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if not rows:
        # create empty file
        with open(path, "w", newline="", encoding="utf-8") as f:
            f.write("")
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        for r in rows:
            # ensure keys exist
            row = {k: ("" if r.get(k) is None else r.get(k)) for k in writer.fieldnames}
            writer.writerow(row)


# ===================================================
# EXCEL EXPORTER (NO DUPLICATE CELLS, CLEAN FORMAT)
# ===================================================
def export_section_grid(rows, timeslots_df, out_path):
    """
    rows: list of dicts or DataFrame-like (section_id, subject_code, teacher_id, room_id, slot_id, day, start_time, end_time)
    timeslots_df: DataFrame of timeslots.csv
    out_path: file path to save workbook
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Side, Font
    import pandas as _pd

    # normalize rows to DataFrame
    if isinstance(rows, list):
        df = _pd.DataFrame(rows)
    else:
        df = rows.copy()

    # ensure required columns
    for c in ["section_id", "subject_code", "teacher_id", "room_id", "slot_id", "day", "start_time", "end_time"]:
        if c not in df.columns:
            df[c] = ""

    wb = Workbook()
    try:
        wb.remove(wb.active)
    except Exception:
        pass

    days_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

    # build ordered list of (day, start_time, slot_id)
    order = []
    for _, r in timeslots_df.iterrows():
        day = r.get("day", "")
        start = r.get("start_time", "")
        sid = r.get("slot_id", "")
        order.append((day, start, sid))
    # filter only days we expect
    order_sorted = sorted(order, key=lambda x: (days_order.index(x[0]) if x[0] in days_order else 99, x[1]))

    # build unique time headers preserving order
    time_labels = []
    slot_for_time = []
    for day, st, sid in order_sorted:
        label = f"{st}-{timeslots_df.loc[timeslots_df['slot_id'] == sid, 'end_time'].iloc[0] if sid in list(timeslots_df['slot_id']) else ''}"
        if label not in time_labels:
            time_labels.append(label)
        slot_for_time.append((day, label, sid))

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Create sheet per section
    sections = sorted(df["section_id"].unique())
    if not sections:
        sections = ["SHEET-EMPTY"]
    for section in sections:
        ws = wb.create_sheet(title=str(section)[:31])
        ws.cell(row=1, column=1, value="DAY / TIME").font = Font(bold=True)

        # write time headers (top row)
        for col_idx, tl in enumerate(time_labels, start=2):
            ws.cell(row=1, column=col_idx, value=tl).font = Font(bold=True)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(14, len(tl) + 2)

        # rows = days
        row_idx = 2
        for d in days_order:
            ws.cell(row=row_idx, column=1, value=d).font = Font(bold=True)
            for col_idx, tl in enumerate(time_labels, start=2):
                # find slot id that matches this day & time label (choose first)
                sid_found = None
                for day, label, sid in slot_for_time:
                    if day == d and label == tl:
                        sid_found = sid
                        break

                if sid_found is None:
                    ws.cell(row=row_idx, column=col_idx, value="").border = border
                    continue

                # get rows for this section and slot
                slot_rows = df[(df["section_id"] == section) & (df["slot_id"] == sid_found)]
                # dedupe combos
                items = set()
                for _, rr in slot_rows.iterrows():
                    subj = rr.get("subject_code", "")
                    teach = rr.get("teacher_id", "")
                    room = rr.get("room_id", "")
                    combo = f"{subj}\n{teach}\n{room}"
                    items.add(combo)
                text = "\n\n".join(sorted(items))
                cell = ws.cell(row=row_idx, column=col_idx, value=text)
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                cell.border = border
            row_idx += 1

    # save
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path


def export_teacher_grid(rows, timeslots_df, out_path):
    """
    rows: list of dicts or DataFrame-like
    timeslots_df: DataFrame
    out_path: path to save
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Side, Font
    import pandas as _pd

    if isinstance(rows, list):
        df = _pd.DataFrame(rows)
    else:
        df = rows.copy()

    for c in ["teacher_id", "section_id", "subject_code", "room_id", "slot_id", "day", "start_time", "end_time"]:
        if c not in df.columns:
            df[c] = ""

    wb = Workbook()
    try:
        wb.remove(wb.active)
    except:
        pass

    days_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

    order = []
    for _, r in timeslots_df.iterrows():
        day = r.get("day", "")
        start = r.get("start_time", "")
        sid = r.get("slot_id", "")
        order.append((day, start, sid))
    order_sorted = sorted(order, key=lambda x: (days_order.index(x[0]) if x[0] in days_order else 99, x[1]))

    time_labels = []
    slot_for_time = []
    for day, st, sid in order_sorted:
        label = f"{st}-{timeslots_df.loc[timeslots_df['slot_id'] == sid, 'end_time'].iloc[0] if sid in list(timeslots_df['slot_id']) else ''}"
        if label not in time_labels:
            time_labels.append(label)
        slot_for_time.append((day, label, sid))

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    teachers = sorted(df["teacher_id"].unique())
    if not teachers:
        teachers = ["T_EMPTY"]
    for teacher in teachers:
        ws = wb.create_sheet(title=str(teacher)[:31])
        ws.cell(row=1, column=1, value="DAY / TIME").font = Font(bold=True)
        for col_idx, tl in enumerate(time_labels, start=2):
            ws.cell(row=1, column=col_idx, value=tl).font = Font(bold=True)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(14, len(tl) + 2)

        row_idx = 2
        for d in days_order:
            ws.cell(row=row_idx, column=1, value=d).font = Font(bold=True)
            for col_idx, tl in enumerate(time_labels, start=2):
                sid_found = None
                for day, label, sid in slot_for_time:
                    if day == d and label == tl:
                        sid_found = sid
                        break

                slot_rows = df[(df["teacher_id"] == teacher) & (df["slot_id"] == sid_found)]
                items = set()
                for _, rr in slot_rows.iterrows():
                    subj = rr.get("subject_code", "")
                    sect = rr.get("section_id", "")
                    room = rr.get("room_id", "")
                    combo = f"{subj} ({sect})\n{room}"
                    items.add(combo)
                text = "\n\n".join(sorted(items))
                cell = ws.cell(row=row_idx, column=col_idx, value=text)
                cell.alignment = Alignment(wrap_text=True)
                cell.border = border
            row_idx += 1

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path


# ==============================================
# TIMETABLE UI CLASS
# ==============================================
class TimetableUI(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("AI Timetable Generator (ORTools + AI Module)")
        self.geometry("1150x720")
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        # auto-detect data folder
        self.project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
        self.data_folder = os.path.join(self.project_root, "data")
        if not os.path.exists(self.data_folder):
            # fallback to cwd if project/data missing
            if os.path.exists(os.path.join(os.getcwd(), "data")):
                self.data_folder = os.path.join(os.getcwd(), "data")
            else:
                self.data_folder = os.getcwd()

        # left frame
        self.left = ctk.CTkFrame(self, width=320)
        self.left.pack(side="left", fill="y", padx=10, pady=10)

        ctk.CTkLabel(self.left, text="AI Timetable Generator", font=("Roboto", 18, "bold")).pack(pady=8)
        ctk.CTkButton(self.left, text="Select CSV Folder", command=self.select_folder).pack(pady=6, fill="x")
        ctk.CTkButton(self.left, text="Generate Timetable (OR-Tools)", fg_color="#008F5D",
                      command=self.run_ortools_thread).pack(pady=8, fill="x")
        ctk.CTkButton(self.left, text="Generate Timetable (Greedy)", fg_color="#444444",
                      command=self.run_greedy_thread).pack(pady=6, fill="x")

        ctk.CTkLabel(self.left, text="AI Command:", anchor="w").pack(pady=(12, 0), padx=6, fill="x")
        self.ai_entry = ctk.CTkEntry(self.left, placeholder_text="e.g. Make TCS-302 double period")
        self.ai_entry.pack(padx=6, pady=6, fill="x")
        ctk.CTkButton(self.left, text="Run AI Command", command=self.run_ai_thread).pack(pady=6, fill="x")

        ctk.CTkButton(self.left, text="Export Grid (Sections)", command=self.export_section_excel).pack(pady=6, fill="x")
        ctk.CTkButton(self.left, text="Export Grid (Teachers)", command=self.export_teacher_excel).pack(pady=6, fill="x")
        ctk.CTkButton(self.left, text="Open timetable.csv", command=self.open_csv).pack(pady=6, fill="x")
        ctk.CTkButton(self.left, text="Quit", fg_color="#8B0000", command=self.destroy).pack(side="bottom", pady=12, padx=8, fill="x")

        # right output
        self.right = ctk.CTkFrame(self)
        self.right.pack(side="right", expand=True, fill="both", padx=8, pady=8)
        self.output = ctk.CTkTextbox(self.right, font=("Consolas", 12), wrap="word")
        self.output.pack(expand=True, fill="both", padx=6, pady=6)

        self.println("Ready. Data folder:", self.data_folder)

    # logging
    def println(self, *msg):
        t = datetime.now().strftime("%H:%M:%S")
        try:
            self.output.insert("end", f"[{t}] {' '.join(map(str, msg))}\n")
            self.output.see("end")
        except Exception:
            print(f"[{t}]", *msg)

    # select folder
    def select_folder(self):
        folder = filedialog.askdirectory(initialdir=self.data_folder)
        if folder:
            self.data_folder = folder
            self.println("Selected folder:", folder)

    # ---------- GREEDY ----------
    def run_greedy_thread(self):
        threading.Thread(target=self.run_greedy, daemon=True).start()

    def run_greedy(self):
        try:
            if scheduler_fix is None:
                self.println("Greedy scheduler module not found.")
                messagebox.showerror("Missing", "scheduler_fix.py not found.")
                return
            self.println("Running Greedy Scheduler...")
            maps = scheduler_fix.build_mappings(self.data_folder)
            assignments, unassigned, slot_ids, slot_meta = scheduler_fix.greedy_schedule(maps)

            rows = []
            for a in assignments:
                sid = a.get("slot_id") or a.get("slot")
                day = ""
                st = ""
                et = ""
                if slot_meta and sid in slot_meta:
                    meta = slot_meta[sid]
                    if isinstance(meta, (list, tuple)) and len(meta) >= 3:
                        day, st, et = meta[0], meta[1], meta[2]
                rows.append({
                    "section_id": a.get("section_id") or a.get("section"),
                    "subject_code": a.get("subject_code") or a.get("subject"),
                    "teacher_id": a.get("teacher_id") or a.get("teacher"),
                    "room_id": a.get("room_id") or a.get("room"),
                    "slot_id": sid,
                    "day": day,
                    "start_time": st,
                    "end_time": et
                })

            _write_csv(os.path.join(self.data_folder, "timetable.csv"), rows)
            self.println("Greedy timetable saved:", os.path.join(self.data_folder, "timetable.csv"))
            if unassigned:
                self.println("⚠ Unassigned:", unassigned)
                messagebox.showwarning("Partial", f"Some items unassigned: {len(unassigned)} (see console).")
            else:
                messagebox.showinfo("Done", "Greedy timetable generated.")
        except Exception as e:
            self.println("Exception during greedy:", e)
            self.println(traceback.format_exc())
            messagebox.showerror("Error", str(e))

    # ---------- ORTOOLS ----------
    def run_ortools_thread(self):
        threading.Thread(target=self.run_ortools, daemon=True).start()

    def run_ortools(self):
        try:
            if ort is None:
                self.println("ORTools scheduler not available.")
                messagebox.showerror("Missing", "scheduler_ortools.py not found.")
                return
            self.println("Running ORTools scheduler...")
            # ort.generate expects base_dir (project root or parent of data)
            base_dir = self.data_folder if os.path.isdir(os.path.join(self.data_folder, "")) else os.getcwd()
            # Many of our ortools wrapper versions expect base (project root) that contains data/
            # If data folder is .../data, pass its parent; otherwise pass data folder itself (code is defensive)
            parent = os.path.dirname(self.data_folder)
            possible_base = parent if os.path.exists(os.path.join(parent, "data")) else self.data_folder
            rows = ort.generate(possible_base)

            if not rows:
                self.println("⚠ ORTools generated 0 rows.")
                messagebox.showwarning("ORTools", "ORTools returned 0 rows (check logs).")
                return

            # rows could be list of dicts (with keys matching expected). Save directly.
            _write_csv(os.path.join(self.data_folder, "timetable.csv"), rows)
            self.println("✔ ORTools timetable saved:", os.path.join(self.data_folder, "timetable.csv"))
            messagebox.showinfo("Done", "ORTools timetable generated.")
        except Exception as e:
            self.println("ORTools exception:", e)
            self.println(traceback.format_exc())
            messagebox.showerror("ORTools Error", str(e))

    # ---------- AI COMMAND ----------
    def run_ai_thread(self):
        threading.Thread(target=self.run_ai, daemon=True).start()

    def run_ai(self):
        cmd = (self.ai_entry.get() or "").strip()
        if not cmd:
            return messagebox.showerror("AI", "Enter an AI command in the box.")
        if updater_ai is None:
            return messagebox.showerror("Missing", "updater_ai.py not available.")
        self.println("AI command received:", cmd)

        def regen():
            self.println("Regenerating timetable via ORTools after AI update...")
            self.run_ortools()

        ok, msg = updater_ai.parse_and_apply(cmd, base_dir=self.data_folder, regen=True, regen_func=regen)
        if ok:
            self.println("AI Success:", msg)
            messagebox.showinfo("AI", msg)
        else:
            self.println("AI Failed:", msg)
            messagebox.showerror("AI Error", msg)

    # ---------- EXPORT SECTION EXCEL ----------
    def export_section_excel(self):
        try:
            tt_path = os.path.join(self.data_folder, "timetable.csv")
            ts_path = os.path.join(self.data_folder, "timeslots.csv")
            if not os.path.exists(tt_path):
                return messagebox.showerror("Missing", "timetable.csv not found. Generate timetable first.")
            if not os.path.exists(ts_path):
                return messagebox.showerror("Missing", "timeslots.csv not found.")

            df_tt = pd.read_csv(tt_path, dtype=str).fillna("")
            df_ts = pd.read_csv(ts_path, dtype=str).fillna("")

            rows = df_tt.to_dict("records")
            out = os.path.join(self.data_folder, "timetable_grid_sections.xlsx")
            export_section_grid(rows, df_ts, out)
            self.println("Section Excel exported:", out)
            messagebox.showinfo("Exported", out)
        except Exception as e:
            self.println("Export error (sections):", e)
            self.println(traceback.format_exc())
            messagebox.showerror("Export Error", str(e))

    # ---------- EXPORT TEACHER EXCEL ----------
    def export_teacher_excel(self):
        try:
            tt_path = os.path.join(self.data_folder, "timetable.csv")
            ts_path = os.path.join(self.data_folder, "timeslots.csv")
            if not os.path.exists(tt_path):
                return messagebox.showerror("Missing", "timetable.csv not found. Generate timetable first.")
            if not os.path.exists(ts_path):
                return messagebox.showerror("Missing", "timeslots.csv not found.")

            df_tt = pd.read_csv(tt_path, dtype=str).fillna("")
            df_ts = pd.read_csv(ts_path, dtype=str).fillna("")

            rows = df_tt.to_dict("records")
            out = os.path.join(self.data_folder, "timetable_grid_teachers.xlsx")
            export_teacher_grid(rows, df_ts, out)
            self.println("Teacher Excel exported:", out)
            messagebox.showinfo("Exported", out)
        except Exception as e:
            self.println("Export error (teachers):", e)
            self.println(traceback.format_exc())
            messagebox.showerror("Export Error", str(e))

    # ---------- OPEN CSV ----------
    def open_csv(self):
        p = os.path.join(self.data_folder, "timetable.csv")
        if os.path.exists(p):
            try:
                os.startfile(p)
            except Exception:
                messagebox.showinfo("Path", p)
        else:
            messagebox.showerror("Missing", "timetable.csv not found. Generate first.")


# run
if __name__ == "__main__":
    app = TimetableUI()
    app.mainloop()
