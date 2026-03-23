# scheduler_fix.py
# Greedy scheduler (kept for debugging / quick schedules)
import argparse, os, sys
import pandas as pd
from collections import defaultdict

def read_csv(path):
    if not os.path.exists(path):
        print("MISSING:", path)
        return None
    return pd.read_csv(path, dtype=str).fillna("")

def build_mappings(data_dir):
    teachers = read_csv(os.path.join(data_dir, "teachers.csv"))
    subjects = read_csv(os.path.join(data_dir, "subjects.csv"))
    tmap = read_csv(os.path.join(data_dir, "teacher_subject_map.csv"))
    ssm = read_csv(os.path.join(data_dir, "section_subject_map.csv"))
    times = read_csv(os.path.join(data_dir, "timeslots.csv"))
    rooms = read_csv(os.path.join(data_dir, "rooms.csv"))
    sections = read_csv(os.path.join(data_dir, "sections.csv"))
    labs = read_csv(os.path.join(data_dir, "labs.csv"))

    missing = []
    for name, df in [("teachers.csv", teachers), ("subjects.csv", subjects),
                     ("teacher_subject_map.csv", tmap), ("section_subject_map.csv", ssm),
                     ("timeslots.csv", times), ("rooms.csv", rooms), ("sections.csv", sections)]:
        if df is None:
            missing.append(name)
    if missing:
        raise RuntimeError("Missing CSVs: " + ", ".join(missing))
    return {
        "teachers": teachers,
        "subjects": subjects,
        "teacher_subject_map": tmap,
        "section_subject_map": ssm,
        "timeslots": times,
        "rooms": rooms,
        "sections": sections,
        "labs": labs
    }

def build_slot_order(times_df):
    day_order = {"Monday":0,"Tuesday":1,"Wednesday":2,"Thursday":3,"Friday":4,"Saturday":5}
    rows = []
    for _, r in times_df.iterrows():
        day = str(r.get("day","")).strip()
        start = str(r.get("start_time","")).strip()
        sid = r.get("slot_id")
        rows.append((day_order.get(day, 99), start, sid))
    rows_sorted = sorted(rows, key=lambda x:(x[0], x[1]))
    slot_ids = [r[2] for r in rows_sorted]
    slot_meta = {}
    for _, r in times_df.iterrows():
        slot_meta[r["slot_id"]] = (r.get("day",""), r.get("start_time",""), r.get("end_time",""))
    return slot_ids, slot_meta

def greedy_schedule(maps):
    times_df = maps["timeslots"]
    slot_ids, slot_meta = build_slot_order(times_df)
    sections_df = maps["sections"]
    ssm_df = maps["section_subject_map"]
    subjects_df = maps["subjects"]
    tmap_df = maps["teacher_subject_map"]
    teachers_df = maps["teachers"]
    rooms_df = maps["rooms"]
    labs_df = maps["labs"]

    teacher_quals = defaultdict(set)
    if tmap_df is not None:
        for _, r in tmap_df.iterrows():
            teacher_quals[r["teacher_id"]].add(r["subject_code"])

    subj_hours = {}
    for _, r in subjects_df.iterrows():
        code = r.get("subject_code")
        hrs = r.get("hours_per_week","")
        try:
            hrs_i = int(float(hrs)) if hrs!="" else 3
        except:
            hrs_i = 3
        subj_hours[code] = max(1, hrs_i)

    section_subjects = defaultdict(list)
    for _, r in ssm_df.iterrows():
        sec = r.get("section_id")
        sc = r.get("subject_code")
        section_subjects[sec].append(sc)

    lab_rooms = []
    lecture_rooms = []
    for _, r in rooms_df.iterrows():
        rid = r.get("room_id")
        typ = r.get("room_type","").lower() if "room_type" in r else ""
        name = r.get("room_name","") if "room_name" in r else ""
        if "lab" in typ or "lab" in rid.lower() or "lab" in name.lower():
            lab_rooms.append(rid)
        else:
            lecture_rooms.append(rid)
    if not lecture_rooms and not lab_rooms:
        lecture_rooms = list(rooms_df["room_id"].unique())

    teacher_busy = defaultdict(set)
    room_busy = defaultdict(set)
    section_busy = defaultdict(set)

    assignments = []
    unassigned = []

    for sec_row in sections_df.itertuples(index=False):
        sec_id = getattr(sec_row, "section_id")
        subjects = section_subjects.get(sec_id, [])
        if not subjects:
            continue
        subs_sorted = sorted(subjects, key=lambda s: -subj_hours.get(s,3))
        for sc in subs_sorted:
            need = subj_hours.get(sc,3)
            placed = 0
            subj_name = subjects_df.loc[subjects_df["subject_code"]==sc, "subject_name"]
            subj_name = subj_name.iloc[0] if not subj_name.empty else ""
            lab_flag = "lab" in str(subj_name).lower() or sc.upper().startswith("PCS")
            cand_teachers = [tid for tid, quals in teacher_quals.items() if sc in quals]
            if not cand_teachers:
                cand_teachers = list(teachers_df["teacher_id"].unique())
            for sid in slot_ids:
                if placed >= need:
                    break
                if sid in section_busy[sec_id]:
                    continue
                rooms_list = lab_rooms if lab_flag and lab_rooms else lecture_rooms
                room_chosen = None
                for r in rooms_list:
                    if sid not in room_busy[r]:
                        room_chosen = r
                        break
                if room_chosen is None:
                    continue
                teacher_chosen = None
                for t in cand_teachers:
                    if sid in teacher_busy[t]:
                        continue
                    teacher_chosen = t
                    break
                if teacher_chosen is None:
                    continue
                assignments.append({
                    "section_id": sec_id,
                    "subject_code": sc,
                    "teacher_id": teacher_chosen,
                    "room_id": room_chosen,
                    "slot_id": sid
                })
                section_busy[sec_id].add(sid)
                room_busy[room_chosen].add(sid)
                teacher_busy[teacher_chosen].add(sid)
                placed += 1
            if placed < need:
                unassigned.append({"section":sec_id, "subject":sc, "needed":need, "placed":placed})
    return assignments, unassigned, slot_ids, slot_meta

def save_outputs(assignments, data_dir, slot_meta):
    import pandas as pd
    df = pd.DataFrame(assignments)
    if df.empty:
        print("No assignments to save.")
        return
    out_csv = os.path.join(data_dir, "timetable.csv")
    df.to_csv(out_csv, index=False)
    print("Saved timetable to", out_csv)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        wb = Workbook()
        wb.remove(wb.active)
        day_slots = {}
        for sid, meta in slot_meta.items():
            day = meta[0]
            day_slots.setdefault(day, []).append((meta[1], sid))
        for d in day_slots:
            day_slots[d] = [sid for _, sid in sorted(day_slots[d])]
        df = pd.DataFrame(assignments)
        for section, group in df.groupby("section_id"):
            ws = wb.create_sheet(title=str(section)[:31])
            times = []
            for d in sorted(day_slots.keys()):
                for sid in day_slots[d]:
                    tlabel = f"{slot_meta[sid][1]}-{slot_meta[sid][2]}"
                    if tlabel not in times:
                        times.append(tlabel)
            ws.cell(row=1, column=1, value="Day / Time").font = Font(bold=True)
            for c, t in enumerate(times, start=2):
                ws.cell(row=1, column=c, value=t).font = Font(bold=True)
            row = 2
            days_sorted = sorted(day_slots.keys(), key=lambda d: ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday"].index(d) if d in ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday"] else 99)
            for d in days_sorted:
                ws.cell(row=row, column=1, value=d).font = Font(bold=True)
                for c, t in enumerate(times, start=2):
                    sid_found = None
                    for sid in day_slots.get(d,[]):
                        label = f"{slot_meta[sid][1]}-{slot_meta[sid][2]}"
                        if label == t:
                            sid_found = sid
                            break
                    if sid_found is None:
                        ws.cell(row=row, column=c, value="")
                        continue
                    rows = group[group["slot_id"]==sid_found]
                    if rows.empty:
                        ws.cell(row=row, column=c, value="")
                    else:
                        texts = []
                        for _, r in rows.iterrows():
                            texts.append(f"{r['subject_code']}\n{r['teacher_id']}\n{r['room_id']}")
                        ws.cell(row=row, column=c, value="\n\n".join(texts))
                        ws.cell(row=row, column=c).alignment = Alignment(wrapText=True)
                row += 1
        out_xlsx = os.path.join(data_dir, "timetable_grid_fixed.xlsx")
        wb.save(out_xlsx)
        print("Saved excel grid to", out_xlsx)
    except Exception as e:
        print("Excel export skipped or failed:", e)

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--data", default="data", help="data folder path")
    args = parser.parse_args()
    data_dir = args.data
    try:
        maps = build_mappings(data_dir)
    except Exception as e:
        print("Error reading CSVs:", e)
        sys.exit(1)
    assignments, unassigned, slot_ids, slot_meta = greedy_schedule(maps)
    print("Assignments made:", len(assignments))
    if unassigned:
        print("UNASSIGNED (section,subject,needed,placed):")
        for u in unassigned:
            print(u)
    save_outputs(assignments, data_dir, slot_meta)
    print("Done.")

if __name__ == "__main__":
    main()
